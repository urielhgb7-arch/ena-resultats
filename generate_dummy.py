import openpyxl

def create_dummy_excel(filename="fictif_import.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes"

    # En-têtes fusionnés (Ligne 1)
    ws.merge_cells('D1:G1')
    ws['D1'] = 'UE: MTH1121'
    
    ws.merge_cells('H1:K1')
    ws['H1'] = 'UE: INF1121'

    # Sous-en-têtes (Ligne 2)
    headers_row2 = ["Numero", "Matricule", "Nom & Prénoms", "EC1", "EC2", "Moy UE", "R", "EC1", "EC2", "Moy UE", "R"]
    for col_num, value in enumerate(headers_row2, 1):
        ws.cell(row=2, column=col_num, value=value)

    # Données étudiants (Ligne 3+)
    students = [
        [1, "MAT001", "KOUASSI JEAN", 14, 15, 14.5, "V", 12, 13, 12.5, "V"],
        [2, "MAT002", "BAMBA ALIOU", 8, 9, 8.5, "NV", 15, 16, 15.5, "V"],
        [3, "MAT003", "KONE AWA", 10, 11, 10.5, "V", 9, 10, 9.5, "NV"],
    ]

    for row_num, student in enumerate(students, 3):
        for col_num, value in enumerate(student, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(filename)
    print(f"Fichier {filename} généré avec succès.")

if __name__ == "__main__":
    create_dummy_excel()
