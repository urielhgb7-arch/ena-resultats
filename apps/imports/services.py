"""
Logique de parsing du fichier Excel source.
Séparée des vues (views.py) pour rester testable sans requête HTTP.

Étape 1 (ce fichier) : lire le fichier Excel "large" (colonnes fusionnées
par UE) et le convertir en lignes brutes stockées telles quelles.
Étape 2 (mapping, à faire une fois le format réel confirmé) : transformer
les LigneBrute en NoteEC / ResultatUE en s'appuyant sur un MappingTemplate.
"""

import openpyxl
from .models import LigneBrute, ImportFichier


def extraire_lignes_brutes(fichier_path, import_fichier: ImportFichier):
    """
    Lit un fichier Excel au format 'large' (en-têtes fusionnés par UE,
    comme les PV réels de l'ENA) et crée une LigneBrute par étudiant,
    sans aucune interprétation métier à ce stade.

    Structure attendue (à confirmer/ajuster une fois le vrai fichier
    du Bureau obtenu) :
      - Ligne d'en-tête 1 : "UE:CODE" fusionné sur plusieurs colonnes
      - Ligne d'en-tête 2 : sous-colonnes (EC1, EC2, Moy UE, R)
      - Lignes suivantes : une ligne par étudiant
    """
    wb = openpyxl.load_workbook(fichier_path, data_only=True)
    ws = wb.active

    # Reconstitue quelle colonne appartient à quelle UE, via les cellules fusionnées
    ue_par_colonne = {}
    for merged_range in ws.merged_cells.ranges:
        valeur = ws.cell(merged_range.min_row, merged_range.min_col).value
        if valeur and "UE" in str(valeur).upper():
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ue_par_colonne[col] = str(valeur).replace("UE:", "").strip()

    lignes_creees = []
    ligne_num = 0

    # Ligne 3 = première ligne étudiant (lignes 1-2 = en-têtes) - à ajuster
    # selon le vrai fichier une fois obtenu.
    for row in ws.iter_rows(min_row=3, values_only=False):
        if row[0].value is None:  # ligne vide = fin du tableau
            continue
        ligne_num += 1

        donnees = {
            "numero": row[0].value,
            "matricule": row[1].value,
            "nom_prenoms": row[2].value,
            "resultats_par_ue": {},
        }

        for col_idx, cell in enumerate(row, start=1):
            if col_idx in ue_par_colonne:
                type_champ = ws.cell(2, col_idx).value  # EC1, EC2, Moy UE, R
                ue_code = ue_par_colonne[col_idx]
                donnees["resultats_par_ue"].setdefault(ue_code, {})
                donnees["resultats_par_ue"][ue_code][str(type_champ)] = cell.value

        ligne = LigneBrute.objects.create(
            import_fichier=import_fichier,
            numero_ligne=ligne_num,
            donnees_brutes=donnees,
        )
        lignes_creees.append(ligne)

    return lignes_creees
